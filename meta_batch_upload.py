"""
Meta 批次上線廣告素材
從【御熹堂】活動檔期_進稿表.xlsx 讀取待上稿列，自動建立 Meta 廣告。

用法：
  python meta_batch_upload.py              # Dry-run（只列出，不實際建立）
  python meta_batch_upload.py --setup      # 列出帳戶下所有廣告活動與廣告組合
  python meta_batch_upload.py --execute    # 實際建立廣告（建立後狀態為 PAUSED）
  python meta_batch_upload.py --execute --activate  # 建立後立即啟用
"""

import argparse
import json
import os
import re
import time
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv()

# ── 設定區（可改為 .env）────────────────────────────────────────────────────────
XLSX_PATH = Path(r"C:\Users\Molly Ho\Downloads\Report\【御熹堂】活動檔期 _ 進稿表.xlsx")
SHEET_NAME = "官網_素材進稿"
GRAPH_URL = "https://graph.facebook.com/v21.0"

META_ACCESS_TOKEN = os.getenv("META_ACCESS_TOKEN", "")
META_AD_ACCOUNT_ID = os.getenv("META_AD_ACCOUNT_ID", "")   # 不含 act_ 前綴
META_PAGE_ID = os.getenv("META_PAGE_ID", "")               # Facebook 粉絲專頁 ID

# ── 欄位對照（1-based）─────────────────────────────────────────────────────────
COL = {
    "時間": 1, "性質": 2, "媒體": 3, "目標": 4, "品項": 5,
    "走期開始": 6, "走期結束": 7, "活動": 8, "料號": 9, "格式": 10,
    "素材": 11, "圖檔": 12, "主標題": 13, "副標題": 14, "文案": 15,
    "CTA連結": 16, "帶品": 17, "法規確認": 18, "可上稿": 19, "已上稿": 20,
    "上稿排程": 21, "素材切角": 22, "廣告名稱": 23, "備註": 24,
    "廣告ID回寫": 25,   # 成功後寫入廣告 ID
}

# ── 品項 → Ad Set ID 對照表（請填入你的實際 ID）──────────────────────────────
# 格式：{ "品項名稱": { "轉換": "adset_id", "流量": "adset_id" } }
# 執行 python meta_batch_upload.py --setup 可列出所有 Campaign / Ad Set
ADSET_MAP: dict[str, dict[str, str]] = {
    "蔓越莓益生菌": {"轉換": "", "流量": ""},
    "魚油":         {"轉換": "", "流量": ""},
    "苦瓜":         {"轉換": "", "流量": ""},
    "豐法":         {"轉換": "", "流量": ""},
    "膠原蛋白":     {"轉換": "", "流量": ""},
    "紅麴":         {"轉換": "", "流量": ""},
    "UC2":          {"轉換": "", "流量": ""},
    "GABA":         {"轉換": "", "流量": ""},
    "健字山苦瓜":   {"轉換": "", "流量": ""},
}

# CTA 按鈕對照
CTA_MAP = {"轉換": "SHOP_NOW", "流量": "LEARN_MORE"}


# ── API helpers ───────────────────────────────────────────────────────────────

def _get(path: str, **params) -> dict:
    r = requests.get(
        f"{GRAPH_URL}/{path}",
        params={"access_token": META_ACCESS_TOKEN, **params},
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()
    if "error" in data:
        raise RuntimeError(data["error"]["message"])
    return data


def _post(path: str, payload: dict) -> dict:
    r = requests.post(
        f"{GRAPH_URL}/{path}",
        json={"access_token": META_ACCESS_TOKEN, **payload},
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    if "error" in data:
        raise RuntimeError(data["error"]["message"])
    return data


# ── 圖片處理 ──────────────────────────────────────────────────────────────────

def gdrive_direct_url(url: str) -> str:
    """把 Google Drive 分享連結轉成可直接下載的 URL。"""
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    return url


def upload_image(url_or_path: str) -> str:
    """上傳圖片至 Meta，回傳 image_hash。"""
    url = gdrive_direct_url(url_or_path)
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()

    content_type = resp.headers.get("Content-Type", "image/jpeg")
    ext = "mp4" if "video" in content_type else "jpg"

    r = requests.post(
        f"{GRAPH_URL}/act_{META_AD_ACCOUNT_ID}/adimages",
        params={"access_token": META_ACCESS_TOKEN},
        files={"filename": (f"creative.{ext}", resp.content, content_type)},
        timeout=120,
    )
    r.raise_for_status()
    data = r.json()
    if "images" in data:
        return next(iter(data["images"].values()))["hash"]
    raise RuntimeError(f"圖片上傳失敗：{data}")


# ── 廣告建立 ──────────────────────────────────────────────────────────────────

def create_creative(image_hash: str, headline: str, body: str,
                    description: str, link: str, cta: str) -> str:
    """建立 Ad Creative，回傳 creative_id。"""
    data = _post(
        f"act_{META_AD_ACCOUNT_ID}/adcreatives",
        {
            "object_story_spec": {
                "page_id": META_PAGE_ID,
                "link_data": {
                    "image_hash": image_hash,
                    "link": link,
                    "message": body,
                    "name": headline,
                    "description": description,
                    "call_to_action": {"type": cta, "value": {"link": link}},
                },
            },
        },
    )
    return data["id"]


def create_ad(adset_id: str, creative_id: str, name: str, activate: bool) -> str:
    """建立廣告，回傳 ad_id。"""
    data = _post(
        f"act_{META_AD_ACCOUNT_ID}/ads",
        {
            "name": name,
            "adset_id": adset_id,
            "creative": {"creative_id": creative_id},
            "status": "ACTIVE" if activate else "PAUSED",
        },
    )
    return data["id"]


# ── 讀取待上稿列 ──────────────────────────────────────────────────────────────

def get_pending(sheet) -> list[tuple[int, dict]]:
    pending = []
    for r in range(2, sheet.max_row + 1):
        can   = str(sheet.cell(r, COL["可上稿"]).value or "").upper()
        done  = str(sheet.cell(r, COL["已上稿"]).value or "").upper()
        media = str(sheet.cell(r, COL["媒體"]).value or "")
        if can == "TRUE" and done != "TRUE" and "Meta" in media:
            pending.append((r, {
                "品項":   str(sheet.cell(r, COL["品項"]).value or "").strip(),
                "目標":   str(sheet.cell(r, COL["目標"]).value or "").strip(),
                "格式":   str(sheet.cell(r, COL["格式"]).value or "").strip(),
                "圖檔":   str(sheet.cell(r, COL["圖檔"]).value or "").strip(),
                "主標題": str(sheet.cell(r, COL["主標題"]).value or "").strip(),
                "副標題": str(sheet.cell(r, COL["副標題"]).value or "").strip(),
                "文案":   str(sheet.cell(r, COL["文案"]).value or "").strip(),
                "CTA連結":str(sheet.cell(r, COL["CTA連結"]).value or "").strip(),
                "廣告名稱":str(sheet.cell(r, COL["廣告名稱"]).value or "").strip(),
            }))
    return pending


# ── Setup：列出帳戶結構 ───────────────────────────────────────────────────────

def cmd_setup():
    """印出所有 Campaign → Ad Set，方便填寫 ADSET_MAP。"""
    if not META_ACCESS_TOKEN or not META_AD_ACCOUNT_ID:
        print("❌ 請先在 .env 設定 META_ACCESS_TOKEN 與 META_AD_ACCOUNT_ID")
        return

    print(f"帳戶：act_{META_AD_ACCOUNT_ID}\n")
    data = _get(
        f"act_{META_AD_ACCOUNT_ID}/campaigns",
        fields="id,name,status",
        limit=50,
    )
    for camp in data.get("data", []):
        print(f"📁 [{camp['status']}] {camp['name']}")
        print(f"   Campaign ID: {camp['id']}")
        adsets = _get(f"{camp['id']}/adsets", fields="id,name,status", limit=100)
        for adset in adsets.get("data", []):
            print(f"   └─ [{adset['status']}] {adset['name']}")
            print(f"      Ad Set ID: {adset['id']}")
        print()


# ── 主流程 ────────────────────────────────────────────────────────────────────

def cmd_run(dry_run: bool, activate: bool):
    if not META_ACCESS_TOKEN or not META_AD_ACCOUNT_ID or not META_PAGE_ID:
        print("❌ 請在 .env 設定 META_ACCESS_TOKEN、META_AD_ACCOUNT_ID、META_PAGE_ID")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    sheet = wb[SHEET_NAME]
    pending = get_pending(sheet)

    print(f"待上稿：{len(pending)} 筆  ({'DRY-RUN' if dry_run else '實際執行'})\n")
    if not pending:
        print("沒有待上稿資料。")
        return

    ok = fail = skip = 0

    for row_idx, d in pending:
        product, goal, fmt = d["品項"], d["目標"], d["格式"]
        ad_name = d["廣告名稱"] or f"{product}_{goal}_{row_idx}"
        print(f"[行 {row_idx}] {ad_name}")

        # 格式過濾：目前只處理單圖，影片/目錄需另行處理
        if fmt not in ("單圖", ""):
            print(f"  ⏭  格式「{fmt}」尚未支援自動上稿，跳過")
            skip += 1
            continue

        # Ad Set ID 查找
        adset_id = ADSET_MAP.get(product, {}).get(goal, "")
        if not adset_id:
            print(f"  ⚠️  ADSET_MAP 缺少 {product!r}/{goal!r}，跳過（執行 --setup 查詢 ID）")
            skip += 1
            continue

        # 必要欄位檢查
        if not d["圖檔"] or not d["CTA連結"]:
            print(f"  ⚠️  缺少圖檔或 CTA 連結，跳過")
            skip += 1
            continue

        if dry_run:
            print(f"  [DRY-RUN] Ad Set: {adset_id}")
            print(f"  圖檔: {d['圖檔'][:80]}")
            print(f"  主標: {d['主標題'][:50]}")
            ok += 1
            continue

        try:
            # 1. 上傳圖片
            image_hash = upload_image(d["圖檔"])
            print(f"  ✅ 圖片 hash: {image_hash}")

            # 2. 建立 Creative
            creative_id = create_creative(
                image_hash=image_hash,
                headline=d["主標題"],
                body=d["文案"],
                description=d["副標題"],
                link=d["CTA連結"],
                cta=CTA_MAP.get(goal, "LEARN_MORE"),
            )
            print(f"  ✅ Creative: {creative_id}")

            # 3. 建立廣告
            ad_id = create_ad(adset_id, creative_id, ad_name, activate)
            print(f"  ✅ 廣告 ID: {ad_id}  狀態: {'ACTIVE' if activate else 'PAUSED'}")

            # 4. 回寫試算表
            sheet.cell(row_idx, COL["已上稿"]).value = True
            sheet.cell(row_idx, COL["廣告ID回寫"]).value = ad_id
            ok += 1

        except Exception as e:
            print(f"  ❌ 失敗：{e}")
            fail += 1

        time.sleep(0.5)  # 避免觸發 rate limit

    if not dry_run and ok > 0:
        wb.save(XLSX_PATH)
        print(f"\n💾 試算表已更新（已上稿欄 + 廣告 ID）")

    print(f"\n{'─'*40}")
    print(f"✅ 成功 {ok}  ❌ 失敗 {fail}  ⏭  跳過 {skip}")


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Meta 批次上線廣告素材")
    parser.add_argument("--setup",    action="store_true", help="列出帳戶所有 Campaign / Ad Set ID")
    parser.add_argument("--execute",  action="store_true", help="實際建立廣告（預設為 dry-run）")
    parser.add_argument("--activate", action="store_true", help="建立後立即設為 ACTIVE（需搭配 --execute）")
    args = parser.parse_args()

    if args.setup:
        cmd_setup()
    else:
        cmd_run(dry_run=not args.execute, activate=args.execute and args.activate)
